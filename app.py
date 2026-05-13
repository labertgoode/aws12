import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# ==========================================
# CONFIGURACIÓN DE LA PÁGINA
# ==========================================
st.set_page_config(page_title="Procesador de Módulos", layout="wide", page_icon="📊")

st.title("📊 Automatización de Módulos de Datos")
st.markdown("Sube tus archivos, procesa la información y descarga los resultados consolidados.")


# ==========================================
# FUNCIONES AUXILIARES
# ==========================================
def to_excel(df):
    """Convierte un DataFrame de Pandas a un archivo Excel en memoria (BytesIO)"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()


# ==========================================
# MENÚ LATERAL Y AYUDA GENERAL
# ==========================================
menu = st.sidebar.selectbox(
    "Selecciona el proceso a ejecutar:",
    ["1. Procesar Lote MI y AI",
     "2. Consolidar Módulo CE",
     "3. Consolidar Módulo B",
     "4. Generar Corte Semanal"]
)

st.sidebar.markdown("---")
with st.sidebar.expander("❓ Ayuda General / Generalidades"):
    st.markdown("""
    **¡Bienvenido al Procesador de Datos!**

    Esta herramienta te ayuda a consolidar y procesar archivos de Excel de forma automática.
    * ⚠️ **Formatos:** Asegúrate de subir siempre archivos en formato `.xlsx`.
    * 🔒 **Privacidad:** Los archivos se procesan en la memoria temporal. Si recargas la página, deberás subirlos de nuevo.
    """)

# ==========================================
# 1. LÓGICA DEL MÓDULO: MI y AI
# ==========================================
if menu == "1. Procesar Lote MI y AI":
    st.header("1. Procesar Lote MI y AI")

    with st.expander("ℹ️ ¿Cómo usar este módulo? (Instrucciones)"):
        st.markdown("""
        **Pasos:**
        1. Sube todos los archivos **Modulo Interactivo** y **Actividad Interactiva** que quieras procesar al mismo tiempo.
        2. Presiona el botón de *Ejecutar Procesamiento*.
        3. Descarga los archivos procesados uno por uno.

        **¿Qué hace este proceso?**
        Busca la columna L (12), revisa si está vacía y rellena la columna de destino correspondiente (13 para MI, 14 para AI) con 'FINALIZADO' o 'NO FINALIZADO'.
        """)

    archivos_mi_ai = st.file_uploader(
        "Sube los archivos MI y AI",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Los archivos que ya tengan '_PROCESADO' en su nombre serán ignorados automáticamente."
    )

    if st.button("Ejecutar Procesamiento", type="primary"):
        if archivos_mi_ai:
            with st.spinner("Procesando archivos..."):
                CONFIG_MODULOS = {"MI": 13, "AI": 14}
                archivos_procesados = []

                for file in archivos_mi_ai:
                    modulo_detectado = next((m for m in CONFIG_MODULOS if m in file.name.upper()), None)

                    if modulo_detectado and "_PROCESADO" not in file.name.upper():
                        col_destino = CONFIG_MODULOS[modulo_detectado]
                        df = pd.read_excel(file)

                        while len(df.columns) <= col_destino:
                            df[f"C_{len(df.columns)}"] = ""


                        def validar(valor):
                            return "NO FINALIZADO" if pd.isna(valor) or str(valor).strip() == "" else "FINALIZADO"


                        df.iloc[:, col_destino] = df.iloc[:, 11].apply(validar)

                        excel_data = to_excel(df)
                        nuevo_nombre = file.name.replace(".xlsx", "_PROCESADO.xlsx")
                        archivos_procesados.append((nuevo_nombre, excel_data))

                if archivos_procesados:
                    st.success(f"✅ Se procesaron {len(archivos_procesados)} archivos con éxito.")
                    for nombre, data in archivos_procesados:
                        st.download_button(label=f"⬇️ Descargar {nombre}", data=data, file_name=nombre,
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    st.warning("No se encontraron archivos válidos MI o AI para procesar.")
        else:
            st.error("Por favor, sube al menos un archivo.")

# ==========================================
# 2. LÓGICA DEL MÓDULO: CE
# ==========================================
elif menu == "2. Consolidar Módulo Ciber Ejercicio":
    st.header("2. Consolidar Módulo Ciber Ejercicio")

    with st.expander("ℹ️ ¿Cómo usar este módulo? (Instrucciones)"):
        st.markdown("""
        **Pasos:**
        1. Sube todos los archivos del módulo **Ciber Ejercicio**.
        2. El sistema limpiará los registros de prueba y unirá todo en un solo archivo.

        **¿Qué hace este proceso?**
        Elimina usuarios de prueba (ej. 'angelica_prueba'), evalúa las columnas L y M para determinar el estatus, une todos los documentos, ordena los datos y elimina duplicados basándose en la primera columna.
        """)

    archivos_ce = st.file_uploader("Sube los archivos CE", type=["xlsx"], accept_multiple_files=True)

    if st.button("Consolidar CE", type="primary"):
        if archivos_ce:
            with st.spinner("Consolidando archivos CE..."):
                lista_dfs = []
                archivos_prueba = []

                for file in archivos_ce:
                    df = pd.read_excel(file)
                    col_usuario = df.columns[0]

                    mascara_prueba = df[col_usuario].astype(str).str.strip().str.lower() == "angelica_prueba"
                    if mascara_prueba.any():
                        archivos_prueba.append(file.name)
                        df = df[~mascara_prueba]

                    while len(df.columns) < 14: df[f"C_{len(df.columns)}"] = ""


                    def evaluar_ce(fila):
                        val_l, val_m = fila.iloc[11], fila.iloc[12]
                        if pd.notna(val_l) and str(val_l).strip() != "":
                            return "FINALIZADO"
                        elif pd.notna(val_m) and str(val_m).strip() != "":
                            return "NO FINALIZADO"
                        return "NO FINALIZADO"


                    df.iloc[:, 13] = df.apply(evaluar_ce, axis=1)
                    lista_dfs.append(df)

                if lista_dfs:
                    df_final = pd.concat(lista_dfs, ignore_index=True)
                    col_n_nombre = df_final.columns[13]
                    df_final = df_final.sort_values(by=col_n_nombre, ascending=True)
                    col_a_nombre = df_final.columns[0]
                    df_final = df_final.drop_duplicates(subset=[col_a_nombre], keep='first')

                    st.success(f"✅ Archivos unidos. Filas totales: {len(df_final)}")
                    if archivos_prueba:
                        st.info(f"🧹 'angelica_prueba' fue eliminada de: {', '.join(archivos_prueba)}")

                    excel_data = to_excel(df_final)
                    st.download_button("⬇️ Descargar CE CONSOLIDADO", data=excel_data,
                                       file_name="CE_CONSOLIDADO_PROCESADO.xlsx")
        else:
            st.error("Por favor, sube al menos un archivo.")

# ==========================================
# 3. LÓGICA DEL MÓDULO: B
# ==========================================
elif menu == "3. Consolidar Módulo Boletin":
    st.header("3. Consolidar Módulo Boletin")

    with st.expander("ℹ️ ¿Cómo usar este módulo? (Instrucciones)"):
        st.markdown("""
        **Pasos:**
        1. Sube todos los archivos del módulo **B**.
        2. El sistema filtrará, evaluará estatus y generará el concentrado.

        **¿Qué hace este proceso?**
        Similar al CE, elimina usuarios de prueba, evalúa la columna K, une todo, elimina filas vacías y quita duplicados asegurando un solo registro por usuario.
        """)

    archivos_b = st.file_uploader("Sube los archivos B", type=["xlsx"], accept_multiple_files=True)

    if st.button("Consolidar B", type="primary"):
        if archivos_b:
            with st.spinner("Consolidando archivos B..."):
                lista_dfs = []
                archivos_prueba = []

                for file in archivos_b:
                    df = pd.read_excel(file)
                    col_usuario = df.columns[0]

                    mascara_prueba = df[col_usuario].astype(str).str.strip().str.lower() == "angelica_prueba"
                    if mascara_prueba.any():
                        archivos_prueba.append(file.name)
                        df = df[~mascara_prueba]

                    while len(df.columns) < 14: df[f"C_{len(df.columns)}"] = ""


                    def evaluar_b(fila):
                        val_k = fila.iloc[10]
                        return "FINALIZADO" if pd.notna(val_k) and str(val_k).strip() != "" else "NO FINALIZADO"


                    df.iloc[:, 13] = df.apply(evaluar_b, axis=1)
                    lista_dfs.append(df)

                if lista_dfs:
                    df_final = pd.concat(lista_dfs, ignore_index=True)
                    col_a_nombre = df_final.columns[0]
                    df_final = df_final.dropna(subset=[col_a_nombre])

                    col_n_nombre = df_final.columns[13]
                    df_final = df_final.sort_values(by=col_n_nombre, ascending=True)

                    df_final['_usuario_temp'] = df_final[col_a_nombre].astype(str).str.strip().str.lower()
                    df_final = df_final.drop_duplicates(subset=['_usuario_temp'], keep='first')
                    df_final = df_final.drop(columns=['_usuario_temp'])

                    st.success(f"✅ Archivos unidos. Usuarios únicos: {len(df_final)}")
                    if archivos_prueba:
                        st.info(f"🧹 'angelica_prueba' fue eliminada de: {', '.join(archivos_prueba)}")

                    excel_data = to_excel(df_final)
                    st.download_button("⬇️ Descargar B CONCENTRADO", data=excel_data,
                                       file_name="B_CONCENTRADO_PROCESADO.xlsx")
        else:
            st.error("Por favor, sube al menos un archivo.")

# ==========================================
# 4. LÓGICA DEL REPORTE FINAL (CORTE SEMANAL)
# ==========================================
elif menu == "4. Generar Corte Semanal":
    st.header("4. Generar Corte Semanal")

    with st.expander("ℹ️ Requisitos e Instrucciones para el Corte Semanal"):
        st.markdown("""
        Para generar el corte, necesitas subir **exactamente 6 archivos** en sus espacios correspondientes:
        * Archivo fuente **MI**
        * Archivo fuente **AI**
        * Archivo fuente **CE**
        * Archivo fuente **B1 (BT)**
        * Archivo fuente **TEST**
        * El **Archivo DESTINO** (Formato maestro).

        📝 **Importante:** Debes escribir el nombre exacto de la pestaña del archivo destino donde se pegarán los datos (Ej: `191125` o `Hoja1`).
        """)

    col1, col2 = st.columns(2)
    with col1:
        archivo_mi = st.file_uploader("1/6 - Archivo MI", type=["xlsx"])
        archivo_ce = st.file_uploader("3/6 - Archivo CE", type=["xlsx"])
        archivo_test = st.file_uploader("5/6 - Archivo TEST", type=["xlsx"])
    with col2:
        archivo_ai = st.file_uploader("2/6 - Archivo AI", type=["xlsx"])
        archivo_bt = st.file_uploader("4/6 - Archivo B1 (BT)", type=["xlsx"])
        archivo_destino = st.file_uploader("6/6 - Archivo DESTINO", type=["xlsx"])

    hoja_dest_name = st.text_input("Nombre de la pestaña de destino (Ej. 191125, Hoja1)", value="Hoja1")

    if st.button("Generar Corte Semanal", type="primary"):
        archivos_list = [archivo_mi, archivo_ai, archivo_ce, archivo_bt, archivo_test, archivo_destino]

        if all(archivos_list) and hoja_dest_name:
            with st.spinner("Ejecutando cruces de información..."):
                try:
                    wb_fuente_MI = load_workbook(BytesIO(archivo_mi.read()), data_only=True)
                    wb_fuente_AI = load_workbook(BytesIO(archivo_ai.read()), data_only=True)
                    wb_fuente_CE = load_workbook(BytesIO(archivo_ce.read()), data_only=True)
                    wb_BT = load_workbook(BytesIO(archivo_bt.read()), data_only=True)
                    wb_TEST = load_workbook(BytesIO(archivo_test.read()), data_only=True)
                    wb_destino = load_workbook(BytesIO(archivo_destino.read()))

                    hoja_MI = wb_fuente_MI["Hoja"] if "Hoja" in wb_fuente_MI.sheetnames else wb_fuente_MI.active
                    hoja_AI = wb_fuente_AI["Hoja"] if "Hoja" in wb_fuente_AI.sheetnames else wb_fuente_AI.active
                    hoja_CE = wb_fuente_CE["Hoja"] if "Hoja" in wb_fuente_CE.sheetnames else wb_fuente_CE.active
                    hoja_BT = wb_BT["CONCENTRADO"] if "CONCENTRADO" in wb_BT.sheetnames else wb_BT.active
                    hoja_TEST = wb_TEST["CENCENTRADO"] if "CENCENTRADO" in wb_TEST.sheetnames else wb_TEST.active

                    if hoja_dest_name not in wb_destino.sheetnames:
                        st.error(f"La hoja '{hoja_dest_name}' no existe en el archivo destino.")
                        st.stop()

                    hoja_destino = wb_destino[hoja_dest_name]
                    ultima_fila_MI = hoja_MI.max_row
                    filas_a_copiar = ultima_fila_MI - 1
                    fila_destino_inicio = 15

                    for i in range(filas_a_copiar):
                        for j in range(8):
                            hoja_destino.cell(row=fila_destino_inicio + i, column=1 + j,
                                              value=hoja_MI.cell(row=2 + i, column=1 + j).value)

                    for i in range(fila_destino_inicio, fila_destino_inicio + filas_a_copiar):
                        celda = hoja_destino.cell(row=i, column=6)
                        if celda.value is not None and isinstance(celda.value, str) and ";" in celda.value:
                            celda.value = celda.value.split(";")[0]

                    for i in range(filas_a_copiar):
                        hoja_destino.cell(row=fila_destino_inicio + i, column=9).value = "FINALIZADO" if hoja_MI.cell(
                            row=2 + i, column=12).value else "NO FINALIZADO"
                        hoja_destino.cell(row=fila_destino_inicio + i, column=10).value = "FINALIZADO" if hoja_AI.cell(
                            row=2 + i, column=12).value else "NO FINALIZADO"
                        hoja_destino.cell(row=fila_destino_inicio + i, column=11, value="FINALIZADO" if (
                                    hoja_CE.cell(row=2 + i, column=12).value or hoja_CE.cell(row=2 + i,
                                                                                             column=13).value) else "NO FINALIZADO")
                        hoja_destino.cell(row=fila_destino_inicio + i, column=12,
                                          value="FINALIZADO" if hoja_BT.cell(row=2 + i,
                                                                             column=11).value else "NO FINALIZADO")
                        hoja_destino.cell(row=fila_destino_inicio + i, column=13, value="FINALIZADO" if (
                                    hoja_TEST.cell(row=2 + i, column=12).value or hoja_TEST.cell(row=2 + i,
                                                                                                 column=13).value) else "NO FINALIZADO")

                    output_corte = BytesIO()
                    wb_destino.save(output_corte)
                    excel_corte_data = output_corte.getvalue()

                    st.success("✅ ¡Corte Semanal generado correctamente!")
                    st.download_button("⬇️ Descargar Corte Semanal", data=excel_corte_data,
                                       file_name=f"CORTE_SEMANAL_{hoja_dest_name}.xlsx")

                except Exception as e:
                    st.error(f"Ocurrió un error procesando los archivos: {e}")
        else:
            st.warning("⚠️ Asegúrate de haber subido los 6 archivos y escrito el nombre de la hoja.")